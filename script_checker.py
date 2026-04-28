import openpyxl
from openpyxl.utils.cell import column_index_from_string

from checks import CHECKS

### There'd better be at least one file here. If you have multiple scripts,
### you can add as many as you have. If you put a simple filename here,
### the file needs to be in the same folder as the python script. Alternatively,
### you can use the full file path here, like "C:\scripts\blah.xlsx"
NAMES = ["common.en.xlsx",
         ]
SHEETS_TO_SKIP = []

# column_index_from_string is 1-indexed, so we need to subtract 1 from the result
JP_COLUMN = column_index_from_string('B') - 1
TL_COLUMN = column_index_from_string('D') - 1
EDIT_COLUMN = column_index_from_string('E') - 1
QA_COLUMN = column_index_from_string('F') - 1

def trim_jp_for_BGI_script_commands(jp):
    return jp.strip(".&")

def format_error_sheet(error_worksheet):
    error_worksheet.title = "Errors"
    error_worksheet.column_dimensions['A'].width = 16
    error_worksheet.column_dimensions['C'].width = 60
    error_worksheet.column_dimensions['D'].width = 60
    
    header = ["Script", "Row", "Line", "Error"]
    error_worksheet.append(header)


def check_line(error_worksheet, script_name, script_row, jp, tl, edit):
    if jp is None:
        return
    
    if tl is None and edit is None and jp != "":
        ### THIS ONE IS CHECKED ALREADY
        ### error_worksheet.append([script_name, script_row, "", "TL and Edit columns both empty?"])
        return
    
    jp = trim_jp_for_BGI_script_commands(jp)
    
    if tl is None:
        tl = edit
    
    tl = tl.strip()
    merge_of_tl_and_edit = tl
    
    if edit is not None and edit != "":
        edit = edit.strip()
        if edit != "":
            merge_of_tl_and_edit = edit
    
    for check in CHECKS:
        error = check(merge_of_tl_and_edit, jp)
        if error is not None:
            error_row = [script_name, script_row, merge_of_tl_and_edit, error]
            error_worksheet.append(error_row)


def main():
    MINROW = 2 # skip the first row with the headers

    error_workbook = openpyxl.Workbook()
    error_worksheet = error_workbook.active
    format_error_sheet(error_worksheet)

    for excel_file_name in NAMES:
        workbook = openpyxl.load_workbook(filename = excel_file_name)
        for sheet_name in workbook.sheetnames:
            if sheet_name in SHEETS_TO_SKIP:
                continue
            sheet = workbook[sheet_name]
            current_row = MINROW
            for row in sheet.iter_rows(min_row=MINROW, values_only=True):
                jp = row[JP_COLUMN]
                tl = row[TL_COLUMN]
                edit = row[EDIT_COLUMN]
                check_line(error_worksheet, sheet_name, current_row, jp, tl, edit)
                current_row += 1

    error_workbook.save("errors_and_warnings.xlsx")
    print("Done.")


main()